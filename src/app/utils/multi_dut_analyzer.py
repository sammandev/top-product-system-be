import csv
import datetime
import io
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.utils.dut_criteria import CriteriaRule, match_rule

# Color fills as requested
FILL_MARGIN_PASS = PatternFill(fill_type="solid", start_color="F8CBAD", end_color="F8CBAD")
FILL_MARGIN_FAIL = PatternFill(fill_type="solid", start_color="FFB1B1", end_color="FFB1B1")
FILL_FAIL = PatternFill(fill_type="solid", start_color="FF0000", end_color="FF0000")


def _read_mc2_df(content: str) -> pd.DataFrame:
    """Read MC2 CSV content into DataFrame. Header is the first row; rows 0=USL,1=LSL,2.. measurements (measurement start always row 4 in file -> df rows index 0..)."""
    df = pd.read_csv(io.StringIO(content), dtype=str, header=0)
    # Ensure empty strings as NaN are avoided
    df = df.fillna("")
    return df


def _normalize_col_for_matching(col: str) -> str:
    x = col.upper()
    # remove WIFI_ prefix if present
    if x.startswith("WIFI_"):
        x = x[5:]
    # replace TX1_,TX2_.. PA1_.. RX1_.. with TX_,PA_,RX_
    x = x.replace("TX1_", "TX_").replace("TX2_", "TX_").replace("TX3_", "TX_").replace("TX4_", "TX_")
    x = x.replace("PA1_", "PA_").replace("PA2_", "PA_").replace("PA3_", "PA_").replace("PA4_", "PA_")
    x = x.replace("RX1_", "RX_").replace("RX2_", "RX_").replace("RX3_", "RX_").replace("RX4_", "RX_")
    return x


def _match_spec_item_to_col(col: str, items: list[str]) -> int:
    """Return index in items that matches column name, or -1."""
    col_norm = _normalize_col_for_matching(col)
    for idx, it in enumerate(items):
        it_norm = it.upper()
        if it_norm in col_norm or col_norm.endswith(it_norm):
            return idx
    return -1


def analyze_mc2_with_spec(
    mc2_content: str,
    spec: dict[str, Any] | None,
    criteria_rules: list[CriteriaRule] | None = None,
) -> tuple[bytes, dict[str, Any]]:
    """Analyze MC2 content and produce compiled XLSX bytes and summary.

    Layout requirements implemented:
    - Two sheets: VALUE_DATA and NON_VALUE_DATA
    - Rows: first a block of metadata rows (Test_Date, Test_Time, ISN, Device_ID, OP_ID, TSP, Script_Ver, Test_Result, TestItem)
    - Then a table where Column1 = Item name, Column2 = USL, Column3 = LSL, Column4.. = measurement per DUT, then one blank column, then comparison columns per DUT (midpoint - measured)
    - Measurement start always at file Row 4 (pandas df rows index 2 onward)
    - Remove any 'PASS' tokens except for metadata fields Test_Result and TestItem
    - Color cells: margin pass F8CBAD, margin fail FFB1B1, fail FF0000, default no fill for PASS
    """
    df = _read_mc2_df(mc2_content)

    # Measurement rows start at index 2 (0-based df rows): header consumed by pandas, then USL (row0), LSL (row1), data starts at row2
    if len(df) < 3:
        raise ValueError("MC2 file does not contain data rows starting at row 4")

    # Skip USL/LSL header rows; data rows start at index 2
    data_rows = df.iloc[2:]

    # Metadata keys to transpose at top (user-requested order).
    # Include ISN so it appears immediately under TestItem in the metadata block.
    meta_keys = [
        "Test_Date",
        "Test_Time",
        "Device_ID",
        "OP_ID",
        "TSP",
        "Script_Ver",
        "Test_Result",
        "TestItem",
        "ISN",
    ]

    # Build list of DUT column labels — we'll use DUT index (1..N) and if available use ISN column values as column headers
    # The data_rows are per DUT; create DUT labels from ISN or Device_ID or row index
    dut_labels = []
    for _, row in data_rows.iterrows():
        label = row.get("ISN") or row.get("Device_ID") or ""
        dut_labels.append(label if label else f"DUT_{len(dut_labels) + 1}")

    # Columns to exclude from the transposed tables (handled separately)
    exclude_cols = set(meta_keys + ["ISN", "Off_Load_Time", "TP_Ver"])

    # Build lists of numeric items (value data) and non-value items
    # We'll scan all columns and if any of their data values in data_rows are numeric we treat as VALUE_DATA else NON_VALUE_DATA
    value_items = []  # tuples of (col_name, list of values)
    nonvalue_items = []

    for col in df.columns:
        # Skip excluded columns which will be handled separately
        if col in exclude_cols:
            continue
        # get data values for data_rows
        vals = [str(v).strip() for v in data_rows[col].tolist()]
        # remove 'PASS' for consideration (we will strip PASS except for metadata)
        vals_no_pass = [v for v in vals if v.upper() != "PASS"]
        is_value = False
        for v in vals_no_pass:
            try:
                float(v)
                is_value = True
                break
            except Exception:
                # hex values like 0x2430 are non-value
                is_value = False
        if is_value:
            value_items.append((col, vals))
        else:
            nonvalue_items.append((col, vals))

    # Prepare workbook
    wb = openpyxl.Workbook()
    ws_val = wb.active
    ws_val.title = "VALUE_DATA"
    ws_non = wb.create_sheet("NON_VALUE_DATA")

    # Helper to write metadata block into a sheet starting at row 1
    def _write_metadata(sheet):
        # metadata key at column 1; reserve columns 2-3 for USL/LSL in table; DUT values start at column 4
        for r_idx, key in enumerate(meta_keys, start=1):
            sheet.cell(row=r_idx, column=1, value=key)
            # write DUT values for this meta key from data_rows starting at col=4
            for i, (_, dr) in enumerate(data_rows.iterrows(), start=1):
                col_idx = 3 + i
                val = dr.get(key, "")
                # keep PASS only for Test_Result and TestItem
                if isinstance(val, str) and val.upper() == "PASS" and key not in ("Test_Result", "TestItem"):
                    val = ""
                sheet.cell(row=r_idx, column=col_idx, value=val)

    # Write metadata to both sheets
    _write_metadata(ws_val)
    _write_metadata(ws_non)

    # After metadata block, set header_row_idx to the first row after metadata (metadata now includes ISN)
    header_row_idx = len(meta_keys) + 1  # immediately after metadata

    # blank column index (after DUT columns)
    blank_col = 3 + len(dut_labels) + 1
    # write DIFF headers in comparison columns (VALUE_DATA only)
    for i, label in enumerate(dut_labels, start=1):
        ws_val.cell(row=header_row_idx, column=blank_col + i, value=f"DIFF_{label}")

    summary = {"total_checked": 0, "total_fail": 0}

    # Helper to get spec arrays for a given section dict
    def _get_spec_lists(metric: dict[str, Any]):
        item_key = next((k for k in metric.keys() if k.endswith("_item")), None)
        if not item_key:
            return [], [], [], None
        base = item_key[:-5]
        items = metric.get(item_key, [])
        max_list = metric.get(f"{base}_max", [])
        min_list = metric.get(f"{base}_min", [])
        gap_list = metric.get(f"{base}_gap", [])
        return items, max_list, min_list, gap_list

    # Build a flattened spec mapping for easier lookup: token -> (usl,lsl,gap,section)
    spec_map: dict[str, tuple[Any, Any, Any, str]] = {}
    sections_present = set()
    if spec:
        for section in ("tx", "pa", "rx", "ibf"):
            for metric in spec.get(section, []):
                enable_key = next((k for k in metric.keys() if k.endswith("_spec_enable")), None)
                if not enable_key or not metric.get(enable_key, False):
                    continue
                items, max_list, min_list, gap_list = _get_spec_lists(metric)
                for idx, item in enumerate(items):
                    usl = None
                    lsl = None
                    gap = None
                    try:
                        if idx < len(max_list) and max_list[idx] not in (False, None, ""):
                            usl = float(max_list[idx])
                    except Exception:
                        usl = None
                    try:
                        if idx < len(min_list) and min_list[idx] not in (False, None, ""):
                            lsl = float(min_list[idx])
                    except Exception:
                        lsl = None
                    try:
                        if gap_list and len(gap_list) > 0:
                            gap = float(gap_list[0])
                    except Exception:
                        gap = None
                    spec_map[item.upper()] = (usl, lsl, gap, section)
                    sections_present.add(section)

    criteria_rules_flat = criteria_rules or []
    if criteria_rules_flat:
        sections_present.add("criteria")

    # Function to find best matching spec for a column name
    def _find_spec_for_col(col_name: str):
        col_norm = _normalize_col_for_matching(col_name)
        for spec_item_upper, (usl, lsl, gap, section) in spec_map.items():
            if spec_item_upper in col_norm:
                return usl, lsl, gap, section
        if criteria_rules_flat:
            rule = match_rule(criteria_rules_flat, col_name)
            if rule is not None:
                usl = rule.usl
                lsl = rule.lsl
                gap = None
                if usl is not None and lsl is not None:
                    gap = abs(usl - lsl) / 4 if usl != lsl else 0.5
                return usl, lsl, gap, "criteria"
        return None, None, None, None

    # Prepare summary counters per DUT and per section
    per_dut_counts: dict[str, dict[str, int]] = {label: {"PASS": 0, "MARGIN_PASS": 0, "MARGIN_FAIL": 0, "FAIL": 0} for label in dut_labels}
    per_section_counts: dict[str, dict[str, int]] = {sec: {"PASS": 0, "MARGIN_PASS": 0, "MARGIN_FAIL": 0, "FAIL": 0} for sec in sections_present}

    # Write VALUE_DATA rows
    # Start immediately after the header_row (no extra blank row)
    start_row = header_row_idx
    for r_offset, (col, vals) in enumerate(value_items):
        row_idx = start_row + r_offset
        ws_val.cell(row=row_idx, column=1, value=col)
        usl, lsl, gap, section = _find_spec_for_col(col)
        if usl is not None:
            ws_val.cell(row=row_idx, column=2, value=usl)
        if lsl is not None:
            ws_val.cell(row=row_idx, column=3, value=lsl)

        # measurement columns
        for i, rawv in enumerate(vals, start=1):
            col_idx = 3 + i
            v = rawv
            # Remove 'PASS' tokens unless metadata
            if isinstance(v, str) and v.upper() == "PASS":
                v = ""
            # try numeric
            try:
                num = float(str(v))
                ws_val.cell(row=row_idx, column=col_idx, value=num)
                # determine status
                status_fill = None
                status_label = "PASS"
                if usl is not None and lsl is not None:
                    if lsl <= num <= usl:
                        status_fill = None  # PASS -> default
                        status_label = "PASS"
                    else:
                        mid = (usl + lsl) / 2.0
                        if gap is not None and abs(num - mid) <= gap:
                            status_fill = FILL_MARGIN_PASS
                            status_label = "MARGIN_PASS"
                        else:
                            # check distance to nearest limit
                            dist_limit = min(abs(num - usl), abs(num - lsl))
                            if gap is not None and dist_limit <= gap:
                                status_fill = FILL_MARGIN_FAIL
                                status_label = "MARGIN_FAIL"
                            else:
                                status_fill = FILL_FAIL
                                status_label = "FAIL"
                # color and count only when we have USL/LSL
                if status_fill is not None:
                    ws_val.cell(row=row_idx, column=col_idx).fill = status_fill
                if usl is not None and lsl is not None:
                    dut_label = dut_labels[i - 1]
                    # increment per-DUT
                    per_dut_counts[dut_label][status_label] += 1
                    # increment per-section
                    if section is not None:
                        per_section_counts.setdefault(
                            section,
                            {"PASS": 0, "MARGIN_PASS": 0, "MARGIN_FAIL": 0, "FAIL": 0},
                        )
                        per_section_counts[section][status_label] += 1
                    # update totals
                    summary["total_checked"] += 1
                    if status_label == "FAIL":
                        summary["total_fail"] += 1
            except Exception:
                # non-numeric slipped to value_items; treat as blank
                ws_val.cell(row=row_idx, column=col_idx, value="")

        # blank col left empty

        # comparison columns: use Excel formula to compute midpoint - measurement
        for i, _rawv in enumerate(vals, start=1):
            comp_col_idx = blank_col + i
            if usl is not None and lsl is not None:
                # measurement cell is at column (3 + i) in this row
                meas_col_idx = 3 + i
                meas_col_letter = get_column_letter(meas_col_idx)
                usl_col_letter = get_column_letter(2)
                lsl_col_letter = get_column_letter(3)
                # Formula: ROUND((USL + LSL)/2 - <meas_cell>, 2)
                formula = f"=ROUND(({usl_col_letter}{row_idx}+{lsl_col_letter}{row_idx})/2 - {meas_col_letter}{row_idx}, 2)"
                ws_val.cell(row=row_idx, column=comp_col_idx, value=formula)
            else:
                ws_val.cell(row=row_idx, column=comp_col_idx, value="")

    # Write NON_VALUE_DATA rows, but skip rows that have no data in any DUT columns (cleaner output)
    # Start immediately after the header_row (no extra blank row)
    start_row_non = header_row_idx
    non_written = 0
    for _, (col, vals) in enumerate(nonvalue_items):
        # check if any DUT value is non-empty (excluding 'PASS')
        has_value = False
        for rawv in vals:
            if isinstance(rawv, str) and rawv.upper() == "PASS":
                continue
            if str(rawv).strip() != "":
                has_value = True
                break
        if not has_value:
            continue

        row_idx = start_row_non + non_written
        ws_non.cell(row=row_idx, column=1, value=col)
        usl, lsl, gap, section = _find_spec_for_col(col)
        if usl is not None:
            ws_non.cell(row=row_idx, column=2, value=usl)
        if lsl is not None:
            ws_non.cell(row=row_idx, column=3, value=lsl)

        for i, rawv in enumerate(vals, start=1):
            col_idx = 3 + i
            v = rawv
            # Keep PASS only for Test_Result/TestItem — these are metadata and not in nonvalue items typically
            if isinstance(v, str) and v.upper() == "PASS":
                v = ""
            ws_non.cell(row=row_idx, column=col_idx, value=v)

        non_written += 1

    # Save workbook to bytes
    # Add SUMMARY sheet
    ws_sum = wb.create_sheet("SUMMARY")
    # Per-DUT table
    ws_sum.cell(row=1, column=1, value="Per-DUT Summary")
    headers = [
        "DUT",
        "PASS",
        "MARGIN_PASS",
        "MARGIN_FAIL",
        "FAIL",
        "TOTAL",
        "PASS_RATE",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws_sum.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True)
    for r, dut in enumerate(dut_labels, start=3):
        counts = per_dut_counts.get(dut, {"PASS": 0, "MARGIN_PASS": 0, "MARGIN_FAIL": 0, "FAIL": 0})
        total = sum(counts.values())
        pass_rate = (counts.get("PASS", 0) / total * 100) if total > 0 else 0.0
        ws_sum.cell(row=r, column=1, value=dut)
        ws_sum.cell(row=r, column=2, value=counts.get("PASS", 0))
        ws_sum.cell(row=r, column=3, value=counts.get("MARGIN_PASS", 0))
        ws_sum.cell(row=r, column=4, value=counts.get("MARGIN_FAIL", 0))
        ws_sum.cell(row=r, column=5, value=counts.get("FAIL", 0))
        ws_sum.cell(row=r, column=6, value=total)
        ws_sum.cell(row=r, column=7, value=round(pass_rate, 2))

    # Adjust column widths for Per-DUT table
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws_sum.column_dimensions[col_letter].width = max(12, len(headers[col_idx - 1]) + 2)

    # Per-section totals (aggregate across DUTs)
    start_sec = 3 + len(dut_labels) + 2
    ws_sum.cell(row=start_sec, column=1, value="Per-Section Summary")
    sec_headers = [
        "Section",
        "PASS",
        "MARGIN_PASS",
        "MARGIN_FAIL",
        "FAIL",
        "TOTAL",
        "PASS_RATE",
    ]
    for c, h in enumerate(sec_headers, start=1):
        cell = ws_sum.cell(row=start_sec + 1, column=c, value=h)
        cell.font = Font(bold=True)
    for idx, (sec, counts) in enumerate(per_section_counts.items(), start=0):
        r = start_sec + 2 + idx
        total = sum(counts.values())
        pass_rate = (counts.get("PASS", 0) / total * 100) if total > 0 else 0.0
        ws_sum.cell(row=r, column=1, value=sec)
        ws_sum.cell(row=r, column=2, value=counts.get("PASS", 0))
        ws_sum.cell(row=r, column=3, value=counts.get("MARGIN_PASS", 0))
        ws_sum.cell(row=r, column=4, value=counts.get("MARGIN_FAIL", 0))
        ws_sum.cell(row=r, column=5, value=counts.get("FAIL", 0))
        ws_sum.cell(row=r, column=6, value=total)
        ws_sum.cell(row=r, column=7, value=round(pass_rate, 2))

    # Adjust column widths for Per-Section table (use same width logic)
    for col_idx in range(1, len(sec_headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws_sum.column_dimensions[col_letter].width = max(12, len(sec_headers[col_idx - 1]) + 2)

    # Also generate a CSV summary in test_outputs
    try:
        out_dir = Path(r"d:\Projects\AST_Parser\backend_fastapi\test_outputs")
        out_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_path = out_dir / f"SUMMARY_{ts}.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as cf:
            writer = csv.writer(cf)
            # write per-DUT header
            writer.writerow(["Per-DUT Summary"])
            writer.writerow(headers)
            for dut in dut_labels:
                counts = per_dut_counts.get(dut, {"PASS": 0, "MARGIN_PASS": 0, "MARGIN_FAIL": 0, "FAIL": 0})
                total = sum(counts.values())
                pass_rate = (counts.get("PASS", 0) / total * 100) if total > 0 else 0.0
                writer.writerow(
                    [
                        dut,
                        counts.get("PASS", 0),
                        counts.get("MARGIN_PASS", 0),
                        counts.get("MARGIN_FAIL", 0),
                        counts.get("FAIL", 0),
                        total,
                        round(pass_rate, 2),
                    ]
                )
            writer.writerow([])
            # per-section
            writer.writerow(["Per-Section Summary"])
            writer.writerow(sec_headers)
            for sec, counts in per_section_counts.items():
                total = sum(counts.values())
                pass_rate = (counts.get("PASS", 0) / total * 100) if total > 0 else 0.0
                writer.writerow(
                    [
                        sec,
                        counts.get("PASS", 0),
                        counts.get("MARGIN_PASS", 0),
                        counts.get("MARGIN_FAIL", 0),
                        counts.get("FAIL", 0),
                        total,
                        round(pass_rate, 2),
                    ]
                )
    except Exception:
        # Fail silently on CSV write (not critical)
        pass

    # Legend with colors
    legend_row = start_sec + 2 + len(per_section_counts) + 2
    ws_sum.cell(row=legend_row, column=1, value="Legend")
    ws_sum.cell(row=legend_row + 1, column=1, value="MARGIN_PASS")
    ws_sum.cell(row=legend_row + 1, column=2, value="")
    ws_sum.cell(row=legend_row + 1, column=2).fill = FILL_MARGIN_PASS
    ws_sum.cell(row=legend_row + 2, column=1, value="MARGIN_FAIL")
    ws_sum.cell(row=legend_row + 2, column=2, value="")
    ws_sum.cell(row=legend_row + 2, column=2).fill = FILL_MARGIN_FAIL
    ws_sum.cell(row=legend_row + 3, column=1, value="FAIL")
    ws_sum.cell(row=legend_row + 3, column=2, value="")
    ws_sum.cell(row=legend_row + 3, column=2).fill = FILL_FAIL

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read(), summary
