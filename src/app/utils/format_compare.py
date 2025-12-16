"""
Utilities to parse MasterControlV2 and WiFi DVT formats and compare metrics.
Functions accept text content (str) so they can be used from CLI or FastAPI uploads.
"""

import csv
import re
from difflib import SequenceMatcher
from io import StringIO
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except Exception:
    Workbook = None
    PatternFill = None
    Font = None

METRICS = ["POW", "EVM", "FREQ", "MASK", "LO_LEAKAGE_DB"]


def format_minimal_number(v):
    """Format number preserving original-like minimal representation.
    - integers shown as '0' or '-5'
    - decimals shown as '21.5' or '-24.5' (no extra trailing zeros)
    - returns 'N/A' for None or non-numeric
    """
    try:
        if v is None:
            return "N/A"
        f = float(v)
        if f.is_integer():
            return f"{int(f)}"
        s = f"{f:.6f}".rstrip("0").rstrip(".")
        return s
    except Exception:
        return "N/A"


def _rows_from_text(text: str):
    f = StringIO(text)
    rdr = csv.reader(f)
    return list(rdr)


def normalize_standard(s: str) -> str:
    if not s:
        return ""
    s = s.strip().upper()
    # canonicalize to keys used in spec_config: 11B, 11AG, 11N, 11AC, 11AX, 11BE
    s = s.replace("A/G", "11AG")
    # map common short tokens to canonical keys
    if s in ("B", "11B"):
        return "11B"
    if s in ("G", "11G"):
        return "11AG"
    if s in ("A", "11A"):
        return "11AG"
    if "11AC" in s or s in ("AC", "11AC"):
        return "11AC"
    if "11AX" in s or s in ("AX", "11AX"):
        return "11AX"
    if s in ("N", "11N"):
        return "11N"
    if "BE" in s or s in ("11BE",):
        return "11BE"
    # fallback: if string contains '11B'/'11G' etc anywhere, prefer that
    if "11B" in s:
        return "11B"
    if "11G" in s:
        return "11AG"
    if "11N" in s:
        return "11N"
    if "11AC" in s:
        return "11AC"
    if "11AX" in s or "AX" in s:
        return "11AX"
    return s


def normalize_datarate(dr: str) -> str:
    if not dr:
        return ""
    d = str(dr).strip().upper()
    # remove stray underscores and extra spaces
    d = d.replace("_", "")
    d = re.sub(r"\s+", "", d)
    # common short forms
    d = d.replace("CCK11", "CCK11")
    d = d.replace("OFDM6", "OFDM6")
    return d


def normalize_bandwidth(bw: str) -> str:
    if not bw:
        return ""
    b = str(bw).strip().upper()
    # common patterns: '20', 'B20', '20MHz', '20 MHZ'
    b = b.replace("MHZ", "")
    b = b.replace(" ", "")
    if b.startswith("B"):
        return b
    if b.isdigit():
        return f"B{b}"
    # fallback
    return b


def parse_mastercontrol_text(
    text: str,
) -> dict[tuple[int, str, int, str, str, str], float]:
    """Parse MasterControlV2 text. Returns mapping keyed by (antenna1, metric, freq, std, datarate, bw)
    e.g. (1, 'POW', 2412, '11B', 'CCK11', 'B20') -> 24.41
    """
    rows = _rows_from_text(text)
    header = None
    data_row = None
    for i, r in enumerate(rows[:8]):
        joined = ",".join(r)
        if re.search(r"\b(TX|RX|PA|RSSI)\d+_", joined, flags=re.IGNORECASE):
            header = r
            # choose next non-empty row
            for j in range(i + 1, min(len(rows), i + 6)):
                if rows[j] and any(cell.strip() for cell in rows[j]):
                    data_row = rows[j]
                    break
            break

    if not header or not data_row:
        raise ValueError("MasterControlV2 header/data row not found")

    # flatten if single field
    if len(data_row) == 1 and "," in data_row[0]:
        values = data_row[0].split(",")
    else:
        values = data_row

    out = {}
    for tok, val in zip(header, values, strict=False):
        tok = tok.strip()
        if not tok:
            continue
        parts = tok.split("_")
        if len(parts) < 3:
            continue
        prefix = parts[0]
        m = re.search(r"(\d+)", prefix)
        if not m:
            continue
        antenna = int(m.group(1))
        # assemble metric name which may be multi-part (e.g., LO_LEAKAGE_DB)
        metric = parts[1].upper()
        remain_idx = 2
        # if the initial metric token is not one of known METRICS, try concatenating
        # subsequent parts until we either match a known metric or run out of tokens
        if metric not in METRICS:
            while remain_idx < len(parts):
                metric = f"{metric}_{parts[remain_idx].upper()}"
                remain_idx += 1
                if metric in METRICS:
                    break
        # if we didn't match a canonical metric, fall back to the original single token
        if metric not in METRICS:
            # reset metric to the single-part token and restart remaining index
            metric = parts[1].upper()
            remain_idx = 2
        # try to extract freq, std, datarate, bw from remaining parts
        freq = None
        std = ""
        dr = ""
        bw = ""
        for p in parts[remain_idx:]:
            if re.match(r"^\d+$", p) and freq is None:
                freq = int(p)
                continue
            # heuristics for standard (contains numbers and letters like 11B, 11N, 11AC)
            if re.match(r"^11[A-Z0-9]+$", p.upper()):
                std = p.upper()
                continue
            # bandwidth like B20, B40, B80
            if re.match(r"^B\d+", p.upper()):
                bw = p.upper()
                continue
            # otherwise treat as datarate
            if not dr:
                dr = p.upper()

        try:
            value = float(val)
        except Exception:
            continue

        out[(antenna, metric, freq, std, dr, bw)] = value

    return out


def parse_wifi_dvt_text(
    text: str,
) -> dict[tuple[int, int, str, str, str], dict[str, Any]]:
    """Parse WiFi DVT text. Returns mapping keyed by (antenna0, freq, std, datarate, bw) -> {metric: value}
    antenna0 is 0-based in DVT.
    """
    rows = _rows_from_text(text)
    title_idx = None
    for i, r in enumerate(rows[:12]):
        # Be tolerant: accept cells that contain the substring 'standard'
        # (some uploads may include BOM or trailing chars)
        if any((cell or "").strip().lower().find("standard") != -1 for cell in r):
            title_idx = i
            break
    if title_idx is None:
        raise ValueError("DVT title row not found")

    # strip BOM from first title if present and trim whitespace
    def _clean_title(t):
        if not t:
            return ""
        s = t.strip()
        # remove common UTF BOM
        if s.startswith("\ufeff"):
            s = s.lstrip("\ufeff")
        return s

    titles = [_clean_title(t) for t in rows[title_idx]]
    data_rows = rows[title_idx + 1 : title_idx + 200]

    out = {}
    for row in data_rows:
        if not any(cell.strip() for cell in row):
            continue
        vals = {titles[i]: (row[i] if i < len(row) else "") for i in range(len(titles))}
        ant_raw = vals.get("Antenna", vals.get("antenna", "0"))
        try:
            antenna = int(ant_raw)
        except Exception:
            antenna = 0

        # frequency field might be 'Freq' or 'Frequency'
        freq_raw = vals.get("Freq", vals.get("Frequency", vals.get("FreqError", "0")))
        try:
            freq = int(str(freq_raw).strip())
        except Exception:
            # try to infer from other columns
            freq = 0

        standard = normalize_standard(vals.get("Standard", ""))
        datarate = normalize_datarate(vals.get("DataRate", vals.get("Data Rate", vals.get("DataRate", ""))))
        bw = normalize_bandwidth(vals.get("BandWidth", vals.get("Bandwidth", vals.get("BandWidth", ""))))

        def getnum(names, vals=vals):
            for n in names:
                if n in vals and vals[n] not in (None, ""):
                    try:
                        return float(vals[n])
                    except Exception:
                        try:
                            return float(str(vals[n]).strip())
                        except Exception:
                            return None
            return None

        entry = {
            "POW": getnum(["M. Power", "M Power", "T. Power", "T Power"]),
            "EVM": getnum(["EVM"]),
            "FREQ": getnum(["FreqError", "Freq Error", "FreqErr Limit", "FreqErr"]),
            "MASK": getnum(["Spectrum Mask", "Mask", "MASK"]),
            "LO_LEAKAGE_DB": getnum(["LO Leakage", "LO Leakage DB", "LO Leakage DB"]),
        }

        out[(antenna, freq, standard, datarate, bw)] = entry
    return out


def _parse_range_token(tok: str):
    """Parse tokens like '10', '-5.5', or '-10~20.5'. Return (min, max) or (None,None) if invalid."""
    if tok is None:
        return (None, None)
    s = str(tok).strip()
    if "~" in s:
        parts = s.split("~")
        try:
            mn = float(parts[0])
            mx = float(parts[1])
            return (mn, mx)
        except Exception:
            return (None, None)
    try:
        v = float(s)
    except Exception:
        return (None, None)
    # single positive number -> symmetric range [-v, v]
    if v >= 0:
        return (-v, v)
    # single negative number -> treat as upper bound (<= v)
    return (None, v)


def _matches_spec(spec_for_std: dict, freq: int, bw: str, dr: str):
    """Find index in spec_for_std that best matches freq/bw/dr. Returns index or None."""
    if not spec_for_std:
        return None
    freqs = spec_for_std.get("frequency", [])
    bws = spec_for_std.get("bw", [])
    mods = spec_for_std.get("mod", [])
    # try to match by frequency first
    sfreq = str(freq)
    for i, f in enumerate(freqs):
        if str(f).strip() == sfreq:
            return i
    # try to match by bandwidth
    for i, b in enumerate(bws):
        if normalize_bandwidth(b) == normalize_bandwidth(bw):
            return i
    # try datarate/mod
    for i, m in enumerate(mods):
        if str(m).strip().upper().replace("_", "") == str(dr).strip().upper().replace("_", ""):
            return i
    return None


def _freq_within_tolerance(m_freq, d_freq, tol_mhz=2.0):
    try:
        if m_freq is None or d_freq is None:
            return False
        return abs(float(m_freq) - float(d_freq)) <= float(tol_mhz)
    except Exception:
        return False


def _datarate_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    a_norm = str(a).upper().replace("_", "").replace(" ", "")
    b_norm = str(b).upper().replace("_", "").replace(" ", "")
    return SequenceMatcher(None, a_norm, b_norm).ratio()


def compute_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Compute aggregated summary per-antenna and per-metric.
    Returns a dict with 'per_antenna' and 'per_metric' aggregates.
    """
    per_antenna = {}
    per_metric = {}
    for r in rows:
        ant = f"Ant_{int(r['antenna_dvt']) + 1}"
        per_antenna.setdefault(ant, {})
        m = r["metric"]
        per_antenna[ant].setdefault("total", 0)
        per_antenna[ant].setdefault("flagged", 0)
        per_antenna[ant]["total"] += 1
        if r.get("flag"):
            per_antenna[ant]["flagged"] += 1

        per_metric.setdefault(m, {"total": 0, "flagged": 0})
        per_metric[m]["total"] += 1
        if r.get("flag"):
            per_metric[m]["flagged"] += 1

    # compute rates
    for v in per_antenna.values():
        total = v.get("total", 0)
        flagged = v.get("flagged", 0)
        v["pass_rate"] = round(((total - flagged) / total) * 100, 2) if total else None

    for v in per_metric.values():
        total = v.get("total", 0)
        flagged = v.get("flagged", 0)
        v["pass_rate"] = round(((total - flagged) / total) * 100, 2) if total else None

    return {"per_antenna": per_antenna, "per_metric": per_metric}


def _check_spec_rule(spec_for_std: dict, index: int, metric: str, value: float) -> bool | None:
    """Return True if value is within spec (i.e., OK), False if out-of-spec, or None if no rule."""
    if not spec_for_std or index is None:
        return None
    # map metric to spec key
    metric_map = {
        "POW": "tx_target_power",
        "EVM": "tx_evm_limit",
        "FREQ": "tx_freq_err_limit",
        "MASK": "tx_mask_limit",
        "LO_LEAKAGE_DB": "tx_lo_leakage_limit",
    }
    key = metric_map.get(metric)
    if not key:
        return None
    vals = spec_for_std.get(key)
    if not vals or index >= len(vals):
        return None
    rule = vals[index]
    if rule is None or str(rule).strip() == "":
        return None
    rule_s = str(rule).strip()
    # exact/target number (positive) -> for POW treat as target, accept small deviation (0.5dB)
    if metric == "POW":
        # rule_s is expected to be the tx_target_power (target) here.
        try:
            target = float(rule_s)
        except Exception:
            return None
        # look for a corresponding tx_target_tolerance list in the provided spec_for_std
        tol_list = spec_for_std.get("tx_target_tolerance") if isinstance(spec_for_std, dict) else None
        # also support a per-standard scalar default 'tx_target_tolerance_default' or
        # scalar 'tx_target_tolerance' (single value) at the spec_for_std level
        std_default_tol = None
        if isinstance(spec_for_std, dict):
            # if spec_for_std provides a scalar 'tx_target_tolerance' treat it as the per-standard default
            if "tx_target_tolerance" in spec_for_std and not isinstance(spec_for_std.get("tx_target_tolerance"), (list, tuple)):
                std_default_tol = spec_for_std.get("tx_target_tolerance")
        tol_min = None
        tol_max = None
        if tol_list and isinstance(tol_list, (list, tuple)) and index < len(tol_list):
            ttok = str(tol_list[index]).strip()
            mn, mx = _parse_range_token(ttok)
            if mn is not None or mx is not None:
                tol_min, tol_max = mn, mx
        # if no per-index list entry, try a per-standard scalar default
        if tol_min is None and tol_max is None and std_default_tol is not None:
            try:
                mn, mx = _parse_range_token(str(std_default_tol))
                if mn is not None or mx is not None:
                    tol_min, tol_max = mn, mx
            except Exception:
                pass
        # If we have an explicit tolerance range, honor it relative to the target
        if tol_min is not None or tol_max is not None:
            # interpret single bounds: if tol_min is None -> value <= target + tol_max
            # if tol_max is None -> value >= target + tol_min
            val = float(value)
            if tol_min is None:
                return val <= (target + tol_max)
            if tol_max is None:
                return val >= (target + tol_min)
            # both present: allowed range is [target + tol_min, target + tol_max]
            return (target + tol_min) <= val <= (target + tol_max)
        # fallback: default symmetric tolerance of 0.5 dB
        try:
            return abs(float(value) - target) <= 0.5
        except Exception:
            return None

    # EVM: spec negative -> value must be <= spec (more negative better)
    if metric == "EVM":
        try:
            evm_limit = float(rule_s)
            return float(value) <= evm_limit
        except Exception:
            return None

    # FREQ: parse range or single numeric -> treat single numeric as symmetric range
    if metric == "FREQ":
        mn, mx = _parse_range_token(rule_s)
        if mn is None and mx is None:
            return None
        val = float(value)
        if mn is None:
            return val <= mx
        if mx is None:
            return val >= mn
        return mn <= val <= mx

    # MASK: expect exact 0 or in-range
    if metric == "MASK":
        if "~" in rule_s:
            mn, mx = _parse_range_token(rule_s)
            if mn is None and mx is None:
                return None
            val = float(value)
            if mn is None:
                return val <= mx
            if mx is None:
                return val >= mn
            return mn <= val <= mx
        try:
            expect = float(rule_s)
            return abs(float(value) - expect) <= 1e-6
        except Exception:
            return None

    # LO_LEAKAGE_DB: ranges
    if metric == "LO_LEAKAGE_DB":
        mn, mx = _parse_range_token(rule_s)
        if mn is None and mx is None:
            return None
        val = float(value)
        if mn is None:
            return val <= mx
        if mx is None:
            return val >= mn
        return mn <= val <= mx

    return None


def compare_maps(
    mc2_map,
    dvt_map,
    threshold: float | None = None,
    spec: dict | None = None,
    freq_tolerance_mhz: float = 2.0,
) -> list[dict[str, Any]]:
    """Compare maps and return a list of comparison dicts, each with metadata.
    Each dict contains: antenna_dvt, antenna_mc2, metric, freq, standard, datarate, bandwidth, mc2_value, dvt_value, diff, flag
    """
    results = []
    # index mc2_map for faster lookups by antenna and metric
    mc2_index = {}
    for (ant, m_metric, m_freq, m_std, m_dr, m_bw), mval in mc2_map.items():
        mc2_index.setdefault((ant, m_metric), []).append(((m_freq, m_std, m_dr, m_bw), mval))

    for (d_ant, d_freq, d_std, d_dr, d_bw), dvals in dvt_map.items():
        mc2_ant = d_ant + 1
        for metric in METRICS:
            dvt_val = dvals.get(metric)
            if dvt_val is None:
                continue

            candidates = mc2_index.get((mc2_ant, metric), [])
            if not candidates:
                continue

            # prefer exact frequency match
            chosen = None
            if d_freq:
                for (m_freq, m_std, m_dr, m_bw), mval in candidates:
                    if m_freq and d_freq and m_freq == d_freq:
                        chosen = ((m_freq, m_std, m_dr, m_bw), mval)
                        break

            # if no exact match, score candidates by frequency closeness (within tolerance) and datarate similarity
            if chosen is None and len(candidates) > 1:
                best = None
                best_score = -1.0
                for (m_freq, m_std, m_dr, m_bw), mval in candidates:
                    score = 0.0
                    # frequency closeness score (1.0 perfect if equal, else decays)
                    try:
                        if m_freq and d_freq:
                            diff = abs(float(m_freq) - float(d_freq))
                            if diff <= freq_tolerance_mhz:
                                score += (
                                    max(
                                        0.0,
                                        (freq_tolerance_mhz - diff) / freq_tolerance_mhz,
                                    )
                                    * 2.0
                                )
                    except Exception:
                        pass
                    # datarate similarity
                    try:
                        score += _datarate_similarity(m_dr or "", d_dr or "")
                    except Exception:
                        pass
                    if score > best_score:
                        best_score = score
                        best = ((m_freq, m_std, m_dr, m_bw), mval)
                if best is not None:
                    chosen = best

            # fallback to first candidate
            if chosen is None:
                chosen = candidates[0]

            (m_freq, m_std, m_dr, m_bw), mval = chosen
            try:
                diff = float(mval) - float(dvt_val)
            except Exception:
                continue

            # determine flag using spec if available, otherwise fall back to threshold (old behavior)
            flag = False
            spec_applied = None
            # per-value spec check results (for MC2 and DVT) so augment_for_human can apply identical logic
            mc2_spec_ok = None
            dvt_spec_ok = None
            # find standard for this comparison (d_std or m_std)
            std_key = (d_std or m_std) or ""
            # try to map to canonical spec key
            if spec and std_key:
                # ensure standard keys are in the same canonical form (e.g., '11B', '11AC')
                canonical_std = normalize_standard(std_key)
                spec_for_std = spec.get(canonical_std) if canonical_std else None
                idx = _matches_spec(
                    spec_for_std,
                    int(d_freq or m_freq or 0),
                    d_bw or m_bw or "",
                    d_dr or m_dr or "",
                )
                # determine which value to use for the metric check: prefer dvt_value if present
                check_value = dvt_val if dvt_val is not None else mval
                try:
                    check_value_f = float(check_value)
                except Exception:
                    check_value_f = None
                if spec_for_std and idx is not None and check_value_f is not None:
                    # handle POW tolerance key if present
                    # the spec may include 'tx_target_tolerance' list aligned with tx_target_power
                    ok = _check_spec_rule(spec_for_std, idx, metric, check_value_f)
                    if ok is not None:
                        # ok indicates the value is within spec for the checked value (dvt preferred)
                        # record both mc2 and dvt checks where appropriate
                        # when checking, ok was computed against 'check_value' which preferred dvt_val
                        # recompute per-value ok for both sides explicitly
                        spec_applied = True
                        try:
                            # check against MC2 value
                            mc2_spec_ok = _check_spec_rule(spec_for_std, idx, metric, float(mval)) if mval is not None else None
                        except Exception:
                            mc2_spec_ok = None
                        try:
                            dvt_spec_ok = _check_spec_rule(spec_for_std, idx, metric, float(dvt_val)) if dvt_val is not None else None
                        except Exception:
                            dvt_spec_ok = None
                        # overall flag should reflect the dvt_val check if available, otherwise mc2
                        if dvt_spec_ok is not None:
                            flag = not bool(dvt_spec_ok)
                        elif mc2_spec_ok is not None:
                            flag = not bool(mc2_spec_ok)
                        else:
                            flag = not bool(ok)
            if spec_applied is None and threshold is not None:
                try:
                    flag = abs(diff) >= float(threshold)
                except Exception:
                    flag = False

            results.append(
                {
                    "antenna_dvt": d_ant,
                    "antenna_mc2": mc2_ant,
                    "metric": metric,
                    "freq": int(d_freq or m_freq) if (d_freq or m_freq) else None,
                    "standard": (d_std or m_std).strip() if (d_std or m_std) else "",
                    "datarate": (d_dr or m_dr).strip() if (d_dr or m_dr) else "",
                    "bandwidth": (d_bw or m_bw).strip() if (d_bw or m_bw) else "",
                    "mc2_value": mval,
                    "dvt_value": dvt_val,
                    "diff": round(diff, 2),
                    "flag": bool(flag),
                    "spec_applied": bool(spec_applied),
                    "mc2_spec_ok": mc2_spec_ok,
                    "dvt_spec_ok": dvt_spec_ok,
                }
            )

    return results


def augment_for_human(
    rows: list[dict[str, Any]],
    spec: dict | None = None,
    runtime_margin_threshold: float | None = None,
) -> list[dict[str, Any]]:
    """Take compare_maps rows and add human-friendly fields:
    - usl, lsl (strings or 'N/A')
    - mc2_spec_diff (formatted pair or single)
    - dvt_spec_diff
    - mc2_dvt_diff (string)
    - result_label (PASS/FAIL/Margin PASS/Margin FAIL)

    This function centralizes formatting so CLI and API are consistent.
    """

    def fmt_num_original(v):
        # Use the module-level minimal formatter for consistency
        return format_minimal_number(v)

    def fmt_pair(left, right):
        # Render left and right preserving minimal formatting
        left_disp = "N/A" if left is None or left == "" else fmt_num_original(left)
        right_disp = "N/A" if right is None or right == "" else fmt_num_original(right)
        return f"{left_disp} | {right_disp}"

    def calc_pair_value(val, usl, lsl):
        # compute (val - usl) and (val - lsl) where applicable, return two floats or None
        if val is None:
            return (None, None)
        try:
            v = float(val)
        except Exception:
            return (None, None)
        left = None
        right = None
        if usl is not None:
            try:
                left = round(v - float(usl), 2)
            except Exception:
                left = None
        if lsl is not None:
            try:
                right = round(v - float(lsl), 2)
            except Exception:
                right = None
        return (left, right)

    def parse_range_token(tok):
        return _parse_range_token(str(tok)) if tok is not None else (None, None)

    out = []
    for r in rows:
        rr = dict(r)
        metric = (rr.get("metric") or "").upper()
        # default displays
        usl = None
        lsl = None
        usl_display = "N/A"
        lsl_display = "N/A"
        mc2_spec_diff = ""
        dvt_spec_diff = ""

        # find spec index for this row
        spec_for_std = None
        idx = None
        if spec and rr.get("standard"):
            canonical = normalize_standard(rr.get("standard"))
            spec_for_std = spec.get(canonical)
            if spec_for_std:
                idx = _matches_spec(
                    spec_for_std,
                    int(rr.get("freq") or 0),
                    rr.get("bandwidth") or "",
                    rr.get("datarate") or "",
                )

        # if spec_for_std exists but idx is None, try to default to the first index
        # when the spec provides per-index lists for the metric (common minimal specs)
        if spec_for_std is not None and idx is None:
            for key in (
                "tx_target_power",
                "tx_evm_limit",
                "tx_freq_err_limit",
                "tx_mask_limit",
                "tx_lo_leakage_limit",
            ):
                vals = spec_for_std.get(key)
                if isinstance(vals, (list, tuple)) and len(vals) > 0:
                    idx = 0
                    break

        mc2_val = rr.get("mc2_value")
        dvt_val = rr.get("dvt_value")

        if metric == "POW":
            if spec_for_std and idx is not None:
                # retrieve target and tolerance
                tp = None
                try:
                    tp = float(spec_for_std.get("tx_target_power")[idx])
                except Exception:
                    tp = None
                tol_val = None
                if "tx_target_tolerance" in spec_for_std:
                    ttv = spec_for_std.get("tx_target_tolerance")
                    if isinstance(ttv, (list, tuple)) and idx < len(ttv):
                        tol_val = ttv[idx]
                    elif not isinstance(ttv, (list, tuple)):
                        tol_val = ttv
                if tol_val is not None:
                    mn, mx = parse_range_token(tol_val)
                else:
                    mn, mx = (-0.5, 0.5)
                # compute usl/lsl relative to target
                if tp is not None:
                    if mn is None and mx is not None:
                        # upper-only
                        usl = tp + mx
                        lsl = None
                    elif mx is None and mn is not None:
                        usl = None
                        lsl = tp + mn
                    else:
                        usl = tp + mx
                        lsl = tp + mn
            else:
                # no spec: leave N/A
                usl = None
                lsl = None

            usl_display = fmt_num_original(usl) if usl is not None else "N/A"
            lsl_display = fmt_num_original(lsl) if lsl is not None else "N/A"

            # For POW: special formatting requirement
            # - MC2 & Spec Diff and DVT & Spec Diff should show comparison to tx_target_power (target)
            #   i.e., value - target (single number) for display.
            # - Internally margin calculation uses differences to USL/LSL (calc_pair_value)
            left_mc2, right_mc2 = calc_pair_value(mc2_val, usl, lsl)
            left_dvt, right_dvt = calc_pair_value(dvt_val, usl, lsl)
            # compute display against target when target exists
            try:
                target = float(spec_for_std.get("tx_target_power")[idx]) if spec_for_std and idx is not None else None
            except Exception:
                target = None
            if target is not None:
                try:
                    mc2_disp = None if mc2_val is None else (float(mc2_val) - target)
                    dvt_disp = None if dvt_val is None else (float(dvt_val) - target)
                    mc2_spec_diff = format_minimal_number(mc2_disp) if mc2_disp is not None else "N/A"
                    dvt_spec_diff = format_minimal_number(dvt_disp) if dvt_disp is not None else "N/A"
                except Exception:
                    mc2_spec_diff = ""
                    dvt_spec_diff = ""
            else:
                mc2_spec_diff = fmt_pair(left_mc2, right_mc2)
                dvt_spec_diff = fmt_pair(left_dvt, right_dvt)

            # compute margin distances using numeric pair values (to USL/LSL)
            def _margin_from_pair_vals(a, b):
                vals = []
                if a is not None:
                    try:
                        vals.append(abs(float(a)))
                    except Exception:
                        pass
                if b is not None:
                    try:
                        vals.append(abs(float(b)))
                    except Exception:
                        pass
                if not vals:
                    return None
                return min(vals)

            md_mc2 = _margin_from_pair_vals(left_mc2, right_mc2)
            md_dvt = _margin_from_pair_vals(left_dvt, right_dvt)

        elif metric in ("EVM", "MASK"):
            key = "tx_evm_limit" if metric == "EVM" else "tx_mask_limit"
            if spec_for_std and idx is not None:
                try:
                    usl_raw = spec_for_std.get(key)[idx]
                    usl = float(usl_raw)
                except Exception:
                    usl = None
            usl_display = fmt_num_original(usl) if usl is not None else "N/A"
            lsl_display = "N/A"
            # single-sided diffs (use minimal formatter)
            try:
                mc2_spec_diff = format_minimal_number(float(mc2_val) - usl) if mc2_val is not None and usl is not None else ("N/A" if mc2_val is not None else "")
            except Exception:
                mc2_spec_diff = ""
            try:
                dvt_spec_diff = format_minimal_number(float(dvt_val) - usl) if dvt_val is not None and usl is not None else ("N/A" if dvt_val is not None else "")
            except Exception:
                dvt_spec_diff = ""

        elif metric in ("FREQ", "LO_LEAKAGE_DB"):
            key = "tx_freq_err_limit" if metric == "FREQ" else "tx_lo_leakage_limit"
            if spec_for_std and idx is not None:
                try:
                    raw = spec_for_std.get(key)[idx]
                    mn, mx = parse_range_token(raw)
                    usl = mx
                    lsl = mn
                except Exception:
                    usl = None
                    lsl = None
            usl_display = fmt_num_original(usl) if usl is not None else "N/A"
            lsl_display = fmt_num_original(lsl) if lsl is not None else "N/A"
            left_mc2, right_mc2 = calc_pair_value(mc2_val, usl, lsl)
            left_dvt, right_dvt = calc_pair_value(dvt_val, usl, lsl)
            # Format with minimal numeric formatting; when missing show 'N/A'
            mc2_spec_diff = fmt_pair(left_mc2, right_mc2)
            dvt_spec_diff = fmt_pair(left_dvt, right_dvt)

        # mc2 vs dvt diff
        try:
            mc2_dvt_diff = ""
            if mc2_val is not None and dvt_val is not None:
                mc2_dvt_diff = format_minimal_number(float(mc2_val) - float(dvt_val))
        except Exception:
            mc2_dvt_diff = ""

        # determine pass/fail and margin
        base_flag = bool(rr.get("flag"))
        result_label = "PASS" if not base_flag else "FAIL"

        # compute margin distance: nearest absolute distance to spec limits
        def margin_distance_from_pair(pair_str):
            try:
                left_s, right_s = [s.strip() for s in str(pair_str).split("|")]
                vals = []
                for s in (left_s, right_s):
                    if s not in ("N/A", ""):
                        try:
                            vals.append(abs(float(s)))
                        except Exception:
                            pass
                if not vals:
                    return None
                return min(vals)
            except Exception:
                return None

        # For non-POW metrics, compute margin distances by parsing the display pair string
        if metric != "POW":
            md_mc2 = margin_distance_from_pair(mc2_spec_diff)
            md_dvt = margin_distance_from_pair(dvt_spec_diff)

        # Determine per-value results (MC2 and DVT) independently based on spec
        def determine_result_for(
            value,
            spec_for_std,
            idx,
            metric,
            md,
            precomputed_ok=None,
            base_flag=base_flag,
        ):
            # value: numeric or None; md: margin distance to nearest limit (abs)
            # spec check
            # If a precomputed spec_ok value is provided (from compare_maps), use it
            if precomputed_ok is not None:
                try:
                    ok = bool(precomputed_ok)
                except Exception:
                    ok = None
            elif spec_for_std and idx is not None and value is not None:
                try:
                    ok = _check_spec_rule(spec_for_std, idx, metric, float(value))
                except Exception:
                    ok = None
            else:
                ok = None

            if ok is None:
                # fallback behavior: if no spec rule available, use base_flag (original flag)
                base = "PASS" if not base_flag else "FAIL"
            else:
                base = "PASS" if ok else "FAIL"
            # determine margin threshold: allow runtime override, then spec to override per-standard or per-metric threshold
            # spec_for_std may include a scalar 'margin_threshold' or metric-specific dict like
            # {'margin_thresholds': {'POW': 0.5, 'EVM': 1.0}}
            margin_threshold = None
            try:
                # runtime override wins
                if runtime_margin_threshold is not None:
                    margin_threshold = float(runtime_margin_threshold)
                else:
                    if spec_for_std and isinstance(spec_for_std, dict):
                        # per-standard scalar
                        if "margin_threshold" in spec_for_std and isinstance(spec_for_std.get("margin_threshold"), (int, float)):
                            margin_threshold = float(spec_for_std.get("margin_threshold"))
                        # metric-specific per-standard overrides
                        mts = spec_for_std.get("margin_thresholds")
                        if margin_threshold is None and isinstance(mts, dict) and metric in mts:
                            try:
                                margin_threshold = float(mts[metric])
                            except Exception:
                                margin_threshold = None
            except Exception:
                margin_threshold = None
            # fall back to default thresholds: POW tighter (0.5 dB), others 1.0 dB
            if margin_threshold is None:
                margin_threshold = 0.5 if metric == "POW" else 1.0
            if md is not None and md <= margin_threshold:
                if base == "PASS":
                    return "M.PASS"
                else:
                    return "M.FAIL"
            return base

        mc2_result = determine_result_for(
            mc2_val,
            spec_for_std,
            idx,
            metric,
            md_mc2,
            precomputed_ok=rr.get("mc2_spec_ok"),
        )
        dvt_result = determine_result_for(
            dvt_val,
            spec_for_std,
            idx,
            metric,
            md_dvt,
            precomputed_ok=rr.get("dvt_spec_ok"),
        )

        # overall result_label: prioritize FAIL, then Margin FAIL, then M. PASS, then PASS
        if mc2_result == "FAIL" or dvt_result == "FAIL":
            result_label = "FAIL"
        elif mc2_result == "M.FAIL" or dvt_result == "M.FAIL":
            result_label = "M.FAIL"
        elif mc2_result == "M.PASS" or dvt_result == "M.PASS":
            result_label = "M.PASS"
        else:
            result_label = "PASS" if not base_flag else "FAIL"

        # attach per-value results
        rr["mc2_result"] = mc2_result
        rr["dvt_result"] = dvt_result

        # attach fields
        rr["usl"] = usl_display
        rr["lsl"] = lsl_display
        rr["mc2_spec_diff"] = mc2_spec_diff
        rr["dvt_spec_diff"] = dvt_spec_diff
        rr["mc2_dvt_diff"] = mc2_dvt_diff
        rr["result_label"] = result_label
        out.append(rr)
    return out


def write_human_xlsx(
    rows: list[dict[str, Any]],
    path,
    provenance: dict | None = None,
    provenance_position: str = "top",
):
    """Write augmented human rows to an XLSX.

    Behavior:
    - Header/order (Antenna, Test Mode, Metric, Freq (no MHz), Standard, DataRate, BW, USL, LSL,
      MC2 Value, MC2 & Spec Diff, MC2 Result, DVT Value, DVT & Spec Diff, DVT Result, MC2 & DVT Diff)
    - Only MC2 Result and DVT Result get colored cell FILLs. Values and SpecDiff cells use font color+bold
      depending on status: PASS -> default, FAIL -> red 'FF0000' bold, M.PASS -> orange 'E26B0A' bold,
      M.FAIL -> dark '963634' bold.
    - M.FAIL fill color changed to 'FFB1B1' for the result fill.
    - Accept either a file path (str/Path) or a writable binary file-like (e.g., BytesIO). If a file-like is
      provided, write to it in-memory and return bytes. If a path string is provided, return default.
    """
    if Workbook is None:
        raise RuntimeError("openpyxl is required to write XLSX files")

    wb = Workbook()
    ws = wb.active
    headers = [
        "Antenna",
        "Mode",
        "Metric",
        "Freq",
        "Standard",
        "DataRate",
        "BW",
        "USL",
        "LSL",
        "MC2 Value",
        "MC2 & Spec Diff",
        "MC2 Result",
        "DVT Value",
        "DVT & Spec Diff",
        "DVT Result",
        "MC2 & DVT Diff",
    ]
    # If provenance provided and position is 'top', write a small metadata row above headers for auditing
    if provenance and provenance_position and provenance_position.lower() in ("top", "header", "above"):
        # write key=value pairs joined by '; '
        meta = "; ".join(f"{k}={v}" for k, v in provenance.items())
        ws.append([meta])
    ws.append(headers)

    # fills for result columns only
    fills = {
        "M.PASS": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        "M.FAIL": PatternFill(start_color="FFB1B1", end_color="FFB1B1", fill_type="solid"),
        "FAIL": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    }

    # font colors for value/spec-diff cells when status indicates change
    font_map = {
        "FAIL": ("FF0000", True),  # red, bold
        "M.PASS": ("E26B0A", True),  # orange, bold
        "M.FAIL": ("963634", True),  # dark maroon, bold
    }

    for r in rows:
        # determine Test Mode
        metric_upper = str(r.get("metric", "")).upper()
        if metric_upper in ("POW", "EVM", "MASK", "FREQ", "LO_LEAKAGE_DB"):
            test_mode = "TX"
        elif metric_upper in ("PER", "RSSI"):
            test_mode = "RX"
        else:
            test_mode = "Others"

        ant_label = int(r.get("antenna_dvt")) + 1 if r.get("antenna_dvt") is not None else ""

        # arrange values to match new header order
        mc2_val_disp = format_minimal_number(r.get("mc2_value")) if r.get("mc2_value") is not None else "N/A"
        dvt_val_disp = format_minimal_number(r.get("dvt_value")) if r.get("dvt_value") is not None else "N/A"

        row_vals = [
            ant_label,
            test_mode,
            r.get("metric"),
            r.get("freq"),
            r.get("standard"),
            r.get("datarate"),
            r.get("bandwidth"),
            r.get("usl", "N/A"),
            r.get("lsl", "N/A"),
            mc2_val_disp,
            r.get("mc2_spec_diff", ""),
            r.get("mc2_result", ""),
            dvt_val_disp,
            r.get("dvt_spec_diff", ""),
            r.get("dvt_result", ""),
            r.get("mc2_dvt_diff", ""),
        ]
        ws.append(row_vals)

        # apply styles according to per-value statuses
        excel_row = ws.max_row
        mc2_result = r.get("mc2_result", "")
        dvt_result = r.get("dvt_result", "")

        # columns (1-based) based on headers
        # Test Mode=1, Antenna=2, Metric=3, Freq=4, Standard=5, DataRate=6, BW=7, USL=8, LSL=9,
        # MC2 Value=10, MC2 & Spec Diff=11, MC2 Result=12, DVT Value=13, DVT & Spec Diff=14, DVT Result=15, MC2 & DVT Diff=16

        # Apply fill only to result columns (12 and 15) if status is not PASS
        def apply_result_fill(status, col, _excel_row=excel_row):
            if not status or status == "PASS":
                return
            fill = fills.get(status)
            if not fill:
                return
            cell = ws.cell(row=_excel_row, column=col)
            cell.fill = fill

        apply_result_fill(mc2_result, 12)
        apply_result_fill(dvt_result, 15)

        # Apply font color+bold to value/spec-diff columns (10,11 for MC2) and (13,14 for DVT)
        def apply_font_for_status(status, cols, _excel_row=excel_row):
            if not status or status == "PASS":
                return
            fcfg = font_map.get(status)
            if not fcfg or Font is None:
                return
            color_hex, is_bold = fcfg
            for c in cols:
                cell = ws.cell(row=_excel_row, column=c)
                cell.font = Font(bold=is_bold, color=color_hex)

        apply_font_for_status(mc2_result, [10, 11])
        apply_font_for_status(dvt_result, [13, 14])

    # If provenance provided and requested at bottom, append as final rows
    if provenance and provenance_position and provenance_position.lower() in ("bottom", "footer", "end"):
        try:
            meta = "; ".join(f"{k}={v}" for k, v in provenance.items())
            # insert an empty row for separation
            ws.append([])
            ws.append([meta])
        except Exception:
            pass

    # If path is a writable binary file-like object, save to it and return bytes
    try:
        # handle BytesIO and other file-likes
        if hasattr(path, "write") and not isinstance(path, (str, bytes)):
            bio = path
            wb.save(bio)
            # seek back to start for caller
            try:
                bio.seek(0)
            except Exception:
                pass
            return bio.getvalue() if hasattr(bio, "getvalue") else None
    except Exception:
        # fallthrough to string path behavior
        pass

    # otherwise assume path is a filesystem path and save
    if isinstance(path, (str,)):
        wb.save(path)
        return None
    # support pathlib.Path
    try:
        from pathlib import Path as _Path

        if isinstance(path, _Path):
            wb.save(str(path))
            return None
    except Exception:
        pass
    # unknown path-like: try saving directly
    wb.save(path)
    return None
