from app.utils.format_compare import write_human_xlsx


def test_write_human_xlsx_creates_file(tmp_path):
    # create a minimal augmented row set
    rows = [
        {
            "antenna_dvt": 0,
            "antenna_mc2": 1,
            "metric": "POW",
            "freq": 2412,
            "standard": "11AC",
            "datarate": "CCK11",
            "bandwidth": "B20",
            "mc2_value": 19.6,
            "dvt_value": 19.4,
            "mc2_spec_diff": "-0.4",
            "dvt_spec_diff": "-0.6",
            "mc2_result": "M.PASS",
            "dvt_result": "M.FAIL",
            "mc2_dvt_diff": "0.2",
            "usl": "20.5",
            "lsl": "19.5",
        }
    ]

    out_file = tmp_path / "human.xlsx"
    # Should not raise - default (top) behavior
    write_human_xlsx(rows, str(out_file))
    # Also ensure bottom-provenance path works
    out_file2 = tmp_path / "human_bottom.xlsx"
    write_human_xlsx(
        rows,
        str(out_file2),
        provenance={"generated": "test"},
        provenance_position="bottom",
    )
    assert out_file.exists()
    # file size should be > 0
    assert out_file.stat().st_size > 0
    # verify the cell fill colors for MC2 Result (col 12) and DVT Result (col 15)
    try:
        from openpyxl import load_workbook
    except Exception:
        return
    wb = load_workbook(str(out_file))
    ws = wb.active
    mc2_res_cell = ws.cell(row=2, column=12)
    dvt_res_cell = ws.cell(row=2, column=15)
    # expected fills mapping
    expected = {
        "M.PASS": "F8CBAD",
        # updated per request: M.FAIL fill color changed to FFB1B1
        "M.FAIL": "FFB1B1",
        "FAIL": "FF0000",
    }

    def _rgb_of(cell):
        f = cell.fill
        if not f:
            return None
        c = None
        try:
            c = f.fgColor.rgb
        except Exception:
            pass
        if not c:
            try:
                c = f.start_color.rgb
            except Exception:
                c = None
        return c

    mc2_fill = _rgb_of(mc2_res_cell)
    dvt_fill = _rgb_of(dvt_res_cell)
    assert mc2_fill is not None and mc2_fill[-6:].upper() == expected["M.PASS"]
    assert dvt_fill is not None and dvt_fill[-6:].upper() == expected["M.FAIL"]
